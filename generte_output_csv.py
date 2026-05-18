import json
import re
from openpyxl import Workbook
from openpyxl.styles import numbers

INPUT_JSON = "/mnt/d/Naved/Outputs/jdcr_derm_vignettes/2022-2025/gpt-5.2-2025-12-11-ablation-r10.json"
OUTPUT_XLSX = "/mnt/d/Naved/Outputs/jdcr_derm_vignettes/2022-2025/xlsx/gpt-5.2-2025-12-11-ablation-r10.xlsx"

# ISSUE_ORDER = [
    # "Dec-25","Jan-24","Feb-24","Mar-24","Apr-24","May-24",
    # "Jun-24","Jul-24","Aug-24","Sep-24","Oct-24","Nov-24",
    # "Dec-24","Jan-25","Feb-25","Mar-25","Apr-25","May-25",
    # "Jun-25","Jul-25","Aug-25","Sep-25","Oct-25","Nov-25",
    # "May-22","Jun-22","Jul-22","Aug-22","Sep-22","Oct-22",
    # "Nov-22","Dec-22","Jan-23","Feb-23","Mar-23","Apr-23",
    # "May-23","Jun-23","Jul-23","Aug-23","Sep-23","Oct-23",
    # "Nov-23","Dec-23"
# ]

ISSUE_ORDER = [
    "May-22", "Jun-22", "Jul-22", "Aug-22", "Sep-22", "Oct-22", "Nov-22", "Dec-22",
    "Jan-23", "Feb-23", "Mar-23", "Apr-23", "May-23", "Jun-23", "Jul-23", "Aug-23",
    "Sep-23", "Oct-23", "Nov-23", "Dec-23", "Jan-24", "Feb-24", "Mar-24", "Apr-24",
    "May-24", "Jun-24", "Jul-24", "Aug-24", "Sep-24", "Oct-24", "Nov-24", "Dec-24",
    "Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25",
    "Sep-25", "Oct-25", "Nov-25", "Dec-25"
]


def normalize_issue(key):
    # 2022Aug → Aug-22
    year = key[:4]
    month = key[4:]
    return f"{month}-{year[-2:]}"

def extract_qnum(text):
    m = re.search(r"Question\s+(\d+)", text)
    return f"Q{m.group(1)}" if m else None

# Load JSON
with open(INPUT_JSON, "r", encoding="utf-8") as f:
    data = json.load(f)

# Build lookup {(issue, Q#): answer_letter}
lookup = {}

for issue_key, questions in data.items():
    issue = normalize_issue(issue_key)
    for q in questions:
        qnum = extract_qnum(q.get("question", ""))
        if qnum:
            lookup[(issue, qnum)] = q.get("answer_letter", "")

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Answers"

# Header
ws.append(["issue", "question", "answer_letter"])

# Force issue column to TEXT
for cell in ws["A"]:
    cell.number_format = numbers.FORMAT_TEXT

# Write rows
row_idx = 2
for issue in ISSUE_ORDER:
    for qnum in ("Q1", "Q2", "Q3"):
        answer = lookup.get((issue, qnum), "")
        if not answer:
            continue        
        ws.cell(row=row_idx, column=1, value=issue)
        ws.cell(row=row_idx, column=2, value=qnum)
        ws.cell(row=row_idx, column=3, value=lookup.get((issue, qnum), ""))

        # Ensure issue column stays TEXT
        ws.cell(row=row_idx, column=1).number_format = numbers.FORMAT_TEXT
        row_idx += 1

# Save file
wb.save(OUTPUT_XLSX)

print(f"Excel file written: {OUTPUT_XLSX}")
print(f"Total rows (excluding header): {row_idx-2}")
